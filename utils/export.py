"""
Export utilities for Factory Management System
Provides Excel export functionality for various modules
"""

import pandas as pd
from io import BytesIO
from flask import make_response
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_response(data, filename, sheet_name="Data"):
    """Create Flask response with Excel file"""
    output = BytesIO()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Add data to worksheet
    for r in dataframe_to_rows(data, index=False, header=True):
        ws.append(r)
    
    # Style the header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # Save workbook
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    
    return response

def export_factory_expenses(expenses):
    """Export factory expenses to Excel"""
    data = []
    for expense in expenses:
        data.append({
            'Expense Number': expense.expense_number,
            'Date': expense.expense_date.strftime('%d/%m/%Y') if expense.expense_date else '',
            'Category': expense.category.replace('_', ' ').title() if expense.category else '',
            'Subcategory': expense.subcategory or '',
            'Description': expense.description,
            'Amount (₹)': float(expense.amount) if expense.amount else 0,
            'Tax Amount (₹)': float(expense.tax_amount) if expense.tax_amount else 0,
            'Total Amount (₹)': float(expense.total_amount) if expense.total_amount else 0,
            'Payment Method': expense.payment_method.replace('_', ' ').title() if expense.payment_method else '',
            'Paid By': expense.paid_by or '',
            'Vendor Name': expense.vendor_name or '',
            'Invoice Number': expense.invoice_number or '',
            'Status': expense.status.title(),
            'Requested By': expense.requested_by.username,
            'Approved By': expense.approved_by.username if expense.approved_by else '',
            'Created Date': expense.created_at.strftime('%d/%m/%Y %H:%M'),
        })
    
    df = pd.DataFrame(data)
    filename = f"factory_expenses_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return create_excel_response(df, filename, "Factory Expenses")

def export_purchase_orders(purchase_orders):
    """Export purchase orders to Excel"""
    data = []
    for po in purchase_orders:
        data.append({
            'PO Number': po.po_number,
            'Date': po.po_date.strftime('%d/%m/%Y') if po.po_date else '',
            'Supplier': po.supplier.name if po.supplier else '',
            'Total Amount (₹)': float(po.total_amount) if po.total_amount else 0,
            'Status': po.status.title(),
            'Delivery Date': po.delivery_date.strftime('%d/%m/%Y') if po.delivery_date else '',
            'Payment Terms': po.payment_terms or '',
            'Notes': po.notes or '',
            'Created By': po.created_by.username if po.created_by else '',
            'Created Date': po.created_at.strftime('%d/%m/%Y %H:%M') if po.created_at else '',
        })
    
    df = pd.DataFrame(data)
    filename = f"purchase_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return create_excel_response(df, filename, "Purchase Orders")

def export_sales_orders(sales_orders):
    """Export sales orders to Excel"""
    data = []
    for so in sales_orders:
        data.append({
            'SO Number': so.so_number,
            'Date': so.so_date.strftime('%d/%m/%Y') if so.so_date else '',
            'Customer': so.customer.name if so.customer else '',
            'Total Amount (₹)': float(so.total_amount) if so.total_amount else 0,
            'Status': so.status.title(),
            'Delivery Date': so.delivery_date.strftime('%d/%m/%Y') if so.delivery_date else '',
            'Payment Terms': so.payment_terms or '',
            'Notes': so.notes or '',
            'Created By': so.created_by.username if so.created_by else '',
            'Created Date': so.created_at.strftime('%d/%m/%Y %H:%M') if so.created_at else '',
        })
    
    df = pd.DataFrame(data)
    filename = f"sales_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return create_excel_response(df, filename, "Sales Orders")

def export_inventory_items(items):
    """Export inventory items to Excel with multi-state inventory details"""
    data = []
    for item in items:
        # Calculate stock values
        stock_value = (item.available_stock or 0) * (item.unit_price or 0)
        
        data.append({
            'Item Code': item.code,
            'Item Name': item.name,
            'Type': item.display_item_type if hasattr(item, 'display_item_type') else (item.item_type.title() if item.item_type else ''),
            'UOM': item.unit_of_measure or '',
            
            # Multi-State Inventory Columns
            'Raw Material': float(item.qty_raw or 0),
            'WIP': float(item.total_wip or 0) if hasattr(item, 'total_wip') else float(item.qty_wip or 0),
            'Finished': float(item.qty_finished or 0),
            'Scrap': float(item.qty_scrap or 0),
            
            # Summary Columns  
            'Total Stock': float(item.total_stock or 0) if hasattr(item, 'total_stock') else float(item.current_stock or 0),
            'Available Stock': float(item.available_stock or 0) if hasattr(item, 'available_stock') else float(item.current_stock or 0),
            'Min Stock': float(item.minimum_stock or 0),
            'Unit Price (₹)': float(item.unit_price) if item.unit_price else 0,
            'Stock Value (₹)': float(stock_value),
            
            # Additional Information
            'Unit Weight (kg)': float(item.unit_weight) if item.unit_weight else 0,
            'HSN Code': item.hsn_code or '',
            'GST Rate (%)': float(item.gst_rate) if item.gst_rate else 0,
            'Last Updated': item.created_at.strftime('%d/%m/%Y') if item.created_at else '',
        })
    
    df = pd.DataFrame(data)
    filename = f"inventory_items_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return create_excel_response(df, filename, "Inventory Items")

def export_employees(employees):
    """Export employees to Excel"""
    data = []
    for emp in employees:
        data.append({
            'Employee Code': emp.employee_code,
            'Name': emp.name,
            'Email': emp.email or '',
            'Phone': emp.phone or '',
            'Department': emp.department or '',
            'Position': emp.position or '',
            'Salary (₹)': float(emp.salary) if emp.salary else 0,
            'Join Date': emp.join_date.strftime('%d/%m/%Y') if emp.join_date else '',
            'Status': 'Active' if emp.is_active else 'Inactive',
            'Address': emp.address or '',
            'Created Date': emp.created_at.strftime('%d/%m/%Y %H:%M') if emp.created_at else '',
        })
    
    df = pd.DataFrame(data)
    filename = f"employees_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return create_excel_response(df, filename, "Employees")

def export_production_orders(production_orders):
    """Export production orders to Excel"""
    data = []
    for prod in production_orders:
        data.append({
            'Production Number': prod.production_number,
            'Item to Produce': prod.item.name if prod.item else '',
            'Planned Quantity': float(prod.planned_quantity) if prod.planned_quantity else 0,
            'Produced Quantity': float(prod.produced_quantity) if prod.produced_quantity else 0,
            'Good Quality': float(prod.good_quality_quantity) if prod.good_quality_quantity else 0,
            'Damaged': float(prod.damaged_quantity) if prod.damaged_quantity else 0,
            'Status': prod.status.title(),
            'Production Date': prod.production_date.strftime('%d/%m/%Y') if prod.production_date else '',
            'Start Date': prod.start_date.strftime('%d/%m/%Y') if prod.start_date else '',
            'End Date': prod.end_date.strftime('%d/%m/%Y') if prod.end_date else '',
            'Notes': prod.notes or '',
            'Created By': prod.created_by.username if prod.created_by else '',
            'Created Date': prod.created_at.strftime('%d/%m/%Y %H:%M') if prod.created_at else '',
        })
    
    df = pd.DataFrame(data)
    filename = f"production_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return create_excel_response(df, filename, "Production Orders")